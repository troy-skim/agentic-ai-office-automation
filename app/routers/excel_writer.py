import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def write_to_excel(data: dict | list[dict], filepath: str, sheet_name: str = "Sheet1"):
    """Write to an Excel file. Create if file doesn't exist.

    Args:
        data (dict | list[dict]): File metadata
        filepath (str): The path to the file
        sheet_name (str, optional): Sheet page title. Defaults to "Sheet1".

    Raises:
        ValueError: _description_
    """
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    if isinstance(data, dict) and "line_items" in data:
        line_items = data.pop("line_items", [])
        invoice_meta = data.copy()  # e.g., invoice_number, date, vendor
        tax_value = invoice_meta.pop("tax", 0.0)  # Remove tax from per-row data

        # Merge invoice metadata with each line item
        flattened = [{**invoice_meta, **item} for item in line_items]
        df = pd.DataFrame(flattened)

        # Write DataFrame to Excel
        df.to_excel(filepath, index=False, sheet_name=sheet_name)

        # Add formula rows
        wb = load_workbook(filepath)
        ws = wb[sheet_name]

        # Excel index
        quantity_col = df.columns.get_loc("quantity") + 1
        unit_price_col = df.columns.get_loc("unit_price") + 1
        last_data_row = ws.max_row + 1

        quantity_letter = get_column_letter(quantity_col)
        unit_price_letter = get_column_letter(unit_price_col)

        # Subtotal
        ws.cell(row=last_data_row, column=unit_price_col - 1, value="Subtotal")
        ws.cell(
            row=last_data_row,
            column=unit_price_col,
            value=f"=SUMPRODUCT({quantity_letter}2:{quantity_letter}{last_data_row - 1}, {unit_price_letter}2:{unit_price_letter}{last_data_row - 1})",
        )

        # Tax
        ws.cell(row=last_data_row + 1, column=unit_price_col - 1, value="Tax")
        ws.cell(row=last_data_row + 1, column=unit_price_col, value=tax_value)

        # Total = subtotal + tax
        ws.cell(row=last_data_row + 2, column=unit_price_col - 1, value="Total")
        ws.cell(
            row=last_data_row + 2,
            column=unit_price_col,
            value=f"={unit_price_letter}{last_data_row}+{unit_price_letter}{last_data_row + 1}",
        )

        wb.save(filepath)
        print(f"📄 Invoice Excel written to: {filepath}")

    elif isinstance(data, dict):
        pd.DataFrame([data]).to_excel(filepath, index=False, sheet_name=sheet_name)
    elif isinstance(data, list):
        pd.DataFrame(data).to_excel(filepath, index=False, sheet_name=sheet_name)
    else:
        raise ValueError("Data must be a dict or list of dicts")
